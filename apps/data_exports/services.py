import csv
from io import BytesIO, StringIO

from apps.employees.models import Employee

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError:  # pragma: no cover
    Workbook = None


EXPORT_COLUMNS = [
    ("id", lambda employee: employee.employee_code or employee.document_number),
    ("area", lambda employee: employee.department),
    ("nombre", lambda employee: " ".join(part for part in [employee.first_name, employee.last_name] if part).strip()),
    ("correo", lambda employee: employee.email),
]


class EmployeeExportService:
    @staticmethod
    def queryset():
        return Employee.objects.all().order_by("last_name", "first_name")

    @staticmethod
    def to_csv() -> bytes:
        buffer = StringIO(newline="")
        writer = csv.writer(buffer, delimiter=";")
        writer.writerow([header for header, _ in EXPORT_COLUMNS])
        for employee in EmployeeExportService.queryset():
            writer.writerow([resolver(employee) for _, resolver in EXPORT_COLUMNS])
        # Excel on Windows detects UTF-8 correctly when the file includes BOM.
        return buffer.getvalue().encode("utf-8-sig")

    @staticmethod
    def to_excel() -> bytes:
        if Workbook is None:
            raise RuntimeError("openpyxl is required to export Excel files.")
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Empleados"
        header_labels = [header.upper() for header, _ in EXPORT_COLUMNS]
        worksheet.append(header_labels)
        for employee in EmployeeExportService.queryset():
            worksheet.append([resolver(employee) for _, resolver in EXPORT_COLUMNS])

        header_fill = PatternFill(fill_type="solid", fgColor="001871")
        header_font = Font(color="FFFFFF", bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")

        for column_index, header in enumerate(header_labels, start=1):
            cell = worksheet.cell(row=1, column=column_index)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

            max_length = len(str(header))
            for row_index in range(2, worksheet.max_row + 1):
                value = worksheet.cell(row=row_index, column=column_index).value
                value_length = len(str(value or ""))
                if value_length > max_length:
                    max_length = value_length

            worksheet.column_dimensions[get_column_letter(column_index)].width = min(max_length + 4, 36)

        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        if worksheet.max_row >= 2 and worksheet.max_column >= 1:
            table = Table(displayName="TablaEmpleados", ref=worksheet.dimensions)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            worksheet.add_table(table)

        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()
