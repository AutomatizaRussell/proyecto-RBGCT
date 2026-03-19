from dataclasses import dataclass
from datetime import date, datetime
from typing import Any

from django.db import transaction
from django.utils import timezone

from apps.common.utils import parse_excel_date
from apps.employees.models import Employee
from apps.integrations.services import IntegrationEventService

from .exceptions import EmployeeImportValidationError
from .models import EmployeeImport, EmployeeImportRow
from .validators import validate_headers

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None


@dataclass
class RowResult:
    status: str
    employee: Employee | None = None
    errors: dict[str, Any] | None = None


class EmployeeImportService:
    STATUS_ALIASES = {
        "activo": Employee.STATUS_ACTIVE,
        "active": Employee.STATUS_ACTIVE,
        "inactivo": Employee.STATUS_INACTIVE,
        "inactive": Employee.STATUS_INACTIVE,
        "suspendido": Employee.STATUS_SUSPENDED,
        "suspended": Employee.STATUS_SUSPENDED,
    }

    FIELD_LABELS = {
        "employee_code": "Código de empleado",
        "document_number": "Número de documento",
        "document_type": "Tipo de documento",
        "first_name": "Nombres",
        "middle_name": "Segundo nombre",
        "last_name": "Apellidos",
        "full_name": "Nombre completo",
        "email": "Correo electrónico",
        "phone": "Teléfono",
        "position": "Cargo",
        "department": "Área o departamento",
        "hire_date": "Fecha de ingreso",
        "birth_date": "Fecha de nacimiento",
        "salary": "Salario",
        "status": "Estado",
    }

    @staticmethod
    def _split_full_name(full_name: str) -> tuple[str, str]:
        cleaned = " ".join(full_name.split())
        if not cleaned:
            return "", ""
        parts = cleaned.split(" ")
        if len(parts) == 1:
            return parts[0], ""
        if len(parts) == 2:
            return parts[0], parts[1]
        midpoint = len(parts) // 2
        return " ".join(parts[:midpoint]), " ".join(parts[midpoint:])

    @staticmethod
    def _normalize_row_payload(row_data: dict[str, Any]) -> dict[str, Any]:
        normalized = dict(row_data)

        employee_identifier = str(
            normalized.get("employee_code") or normalized.get("document_number") or ""
        ).strip()
        if employee_identifier:
            if not str(normalized.get("employee_code", "")).strip():
                normalized["employee_code"] = employee_identifier
            if not str(normalized.get("document_number", "")).strip():
                normalized["document_number"] = employee_identifier

        if normalized.get("middle_name") and not normalized.get("first_name"):
            normalized["first_name"] = normalized.get("middle_name")

        full_name = str(normalized.get("full_name", "")).strip()
        if full_name and (not str(normalized.get("first_name", "")).strip() or not str(normalized.get("last_name", "")).strip()):
            first_name, last_name = EmployeeImportService._split_full_name(full_name)
            normalized["first_name"] = str(normalized.get("first_name") or first_name).strip()
            normalized["last_name"] = str(normalized.get("last_name") or last_name).strip()

        normalized["first_name"] = str(normalized.get("first_name", "")).strip()
        normalized["last_name"] = str(normalized.get("last_name", "")).strip()

        status_value = str(normalized.get("status", "")).strip().lower()
        if status_value:
            normalized["status"] = EmployeeImportService.STATUS_ALIASES.get(status_value, status_value)

        return normalized

    @staticmethod
    def mark_failed(import_batch: EmployeeImport, *, message: str, details: dict[str, Any] | None = None) -> EmployeeImport:
        import_batch.status = EmployeeImport.STATUS_FAILED
        import_batch.finished_at = timezone.now()
        import_batch.summary = {
            "error": message,
            "details": details or {},
        }
        import_batch.save(update_fields=["status", "finished_at", "summary", "updated_at"])
        return import_batch

    @staticmethod
    def _normalize_json_value(value):
        if isinstance(value, (datetime, date)):
            return value.isoformat()
        if isinstance(value, dict):
            return {key: EmployeeImportService._normalize_json_value(inner) for key, inner in value.items()}
        if isinstance(value, (list, tuple)):
            return [EmployeeImportService._normalize_json_value(inner) for inner in value]
        return value

    @staticmethod
    def process(import_batch: EmployeeImport) -> EmployeeImport:
        if load_workbook is None:
            raise RuntimeError("openpyxl is required to process Excel files.")

        import_batch.status = EmployeeImport.STATUS_PROCESSING
        import_batch.started_at = timezone.now()
        import_batch.save(update_fields=["status", "started_at", "updated_at"])

        try:
            workbook = load_workbook(import_batch.file.path)
            worksheet = workbook.active
            header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
            headers = validate_headers(list(header_row))
        except StopIteration as exc:
            EmployeeImportService.mark_failed(
                import_batch,
                message="El archivo Excel esta vacio o no contiene encabezados.",
            )
            raise EmployeeImportValidationError("El archivo Excel esta vacio o no contiene encabezados.") from exc
        except EmployeeImportValidationError as exc:
            EmployeeImportService.mark_failed(
                import_batch,
                message=exc.message,
                details={"missing_columns": exc.missing_columns},
            )
            raise
        except Exception as exc:
            EmployeeImportService.mark_failed(
                import_batch,
                message="No fue posible leer el archivo Excel.",
                details={"exception": str(exc)},
            )
            raise EmployeeImportValidationError("No fue posible leer el archivo Excel. Verifica que sea un .xlsx valido.") from exc

        total_rows = 0
        success_rows = 0
        error_rows = 0

        for index, values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(values):
                continue
            total_rows += 1
            payload = EmployeeImportService._normalize_row_payload(
                {header: value for header, value in zip(headers, values) if header}
            )
            result = EmployeeImportService._process_row(payload)

            EmployeeImportRow.objects.update_or_create(
                import_batch=import_batch,
                row_number=index,
                defaults={
                    "raw_data": EmployeeImportService._normalize_json_value(payload),
                    "status": result.status,
                    "error_detail": result.errors or {},
                    "employee": result.employee,
                },
            )

            if result.status == EmployeeImportRow.STATUS_SUCCESS:
                success_rows += 1
            else:
                error_rows += 1

        import_batch.total_rows = total_rows
        import_batch.processed_rows = success_rows + error_rows
        import_batch.success_rows = success_rows
        import_batch.error_rows = error_rows
        import_batch.finished_at = timezone.now()
        import_batch.status = (
            EmployeeImport.STATUS_COMPLETED
            if error_rows == 0
            else EmployeeImport.STATUS_PARTIAL if success_rows > 0 else EmployeeImport.STATUS_FAILED
        )
        import_batch.summary = {
            "total_rows": total_rows,
            "success_rows": success_rows,
            "error_rows": error_rows,
        }
        import_batch.save()

        IntegrationEventService.queue_event(
            event_type="employee_import.completed",
            payload={
                "import_id": import_batch.id,
                "status": import_batch.status,
                "summary": import_batch.summary,
            },
        )
        return import_batch

    @staticmethod
    @transaction.atomic
    def _process_row(row_data: dict[str, Any]) -> RowResult:
        required_fields = ["employee_code", "first_name", "document_number"]
        missing_fields = [field for field in required_fields if not row_data.get(field)]
        if missing_fields:
            return RowResult(
                status=EmployeeImportRow.STATUS_ERROR,
                errors={"missing_fields": missing_fields},
            )

        status_value = str(row_data.get("status", Employee.STATUS_ACTIVE)).strip() or Employee.STATUS_ACTIVE
        if status_value not in {choice[0] for choice in Employee.STATUS_CHOICES}:
            return RowResult(
                status=EmployeeImportRow.STATUS_ERROR,
                errors={"status": f"Estado no válido: '{status_value}'."},
            )

        defaults = {
            "first_name": str(row_data.get("first_name", "")).strip(),
            "last_name": str(row_data.get("last_name", "")).strip(),
            "document_type": str(row_data.get("document_type", "")).strip(),
            "email": str(row_data.get("email", "")).strip(),
            "phone": str(row_data.get("phone", "")).strip(),
            "position": str(row_data.get("position", "")).strip(),
            "department": str(row_data.get("department", "")).strip(),
            "hire_date": parse_excel_date(row_data.get("hire_date")),
            "birth_date": parse_excel_date(row_data.get("birth_date")),
            "salary": row_data.get("salary") or None,
            "status": status_value,
            "metadata": {
                "source": "excel_import",
                "raw": EmployeeImportService._normalize_json_value(
                    {k: v for k, v in row_data.items() if k not in required_fields}
                ),
            },
        }

        employee, _ = Employee.objects.update_or_create(
            document_number=str(row_data["document_number"]).strip(),
            defaults={
                "employee_code": str(row_data["employee_code"]).strip(),
                **defaults,
            },
        )
        return RowResult(status=EmployeeImportRow.STATUS_SUCCESS, employee=employee)

    @classmethod
    def get_supported_columns(cls) -> list[dict[str, str]]:
        return [
            {"field": "employee_code", "label": cls.FIELD_LABELS["employee_code"], "aliases": "codigo_empleado, codigo, id"},
            {"field": "first_name", "label": cls.FIELD_LABELS["first_name"], "aliases": "nombres, primer_nombre"},
            {"field": "last_name", "label": cls.FIELD_LABELS["last_name"], "aliases": "apellidos, apellido, primer_apellido"},
            {"field": "full_name", "label": cls.FIELD_LABELS["full_name"], "aliases": "nombre, nombre_completo, completo"},
            {"field": "document_number", "label": cls.FIELD_LABELS["document_number"], "aliases": "numero_documento, nro_documento, documento"},
            {"field": "document_type", "label": cls.FIELD_LABELS["document_type"], "aliases": "tipo_documento"},
            {"field": "email", "label": cls.FIELD_LABELS["email"], "aliases": "correo, correo_electronico"},
            {"field": "phone", "label": cls.FIELD_LABELS["phone"], "aliases": "telefono, celular"},
            {"field": "position", "label": cls.FIELD_LABELS["position"], "aliases": "cargo, puesto"},
            {"field": "department", "label": cls.FIELD_LABELS["department"], "aliases": "departamento, area"},
            {"field": "hire_date", "label": cls.FIELD_LABELS["hire_date"], "aliases": "fecha_ingreso, ingreso"},
            {"field": "birth_date", "label": cls.FIELD_LABELS["birth_date"], "aliases": "fecha_nacimiento, nacimiento"},
            {"field": "salary", "label": cls.FIELD_LABELS["salary"], "aliases": "salario"},
            {"field": "status", "label": cls.FIELD_LABELS["status"], "aliases": "estado"},
        ]
